from selenium import webdriver
from selenium.webdriver.common import keys
from selenium.webdriver.common.keys import Keys
from time import sleep, strftime
from random import randint
import pandas as pd
from openpyxl import load_workbook
import os
import sys

dirname = 'D:\\Report\\1400\\TabdilVaz'
progdirname = os.path.dirname(__file__)
       
chromedriver_path = r'D://chromedriver/chromedriver.exe' # Change this to your own chromedriver path!
webdriver = webdriver.Chrome(executable_path=chromedriver_path)
# webdriver.set_window_size(1920, 1080)
webdriver.maximize_window()
webdriver.get('https://amar.imo.org.ir/')
html = webdriver.page_source
userData = open("userData.udata", "r")

username = webdriver.find_element_by_name('_58_INSTANCE_dehyari_login')
username.send_keys(userData.readline().rstrip())
password = webdriver.find_element_by_name('_58_INSTANCE_dehyari_password')
password.send_keys(userData.readline().rstrip())

webdriver.find_element_by_xpath('//button[text()=" ورود "]').click()
companies = userData.readline().rstrip().split(',')
print(companies)
# companies = ['SazMotori','SazMotori']
forms =  userData.readline().rstrip().split(',')
excelfile =  userData.readline().rstrip()
if len(sys.argv) >= 2:
    excelfile = sys.argv[1]
if len(sys.argv) >= 3:
    forms = sys.argv[2].split(',')
print(forms)
i = 0
delaytime = 3
errorMsg = ''


# داشبورد مدیریت
webdriver.get('https://amar.imo.org.ir/member/-_-irn-l-14000277232')
webdriver.find_element_by_xpath("//*[contains(text(),'نمایش داده های آماری')]/..").click()
sleep(delaytime)
mainFrame = webdriver.find_element_by_xpath("//iframe")
# webdriver.switch_to_frame('_cartable_WAR_Cartableportlet_dashboard-application_iframe_')
webdriver.switch_to.frame(mainFrame)
webdriver.find_element_by_xpath("//*[contains(text(),'مجموعه اطلاعات جامع اداری و استخدامی')]").click()
webdriver.find_element_by_xpath("//*[contains(text(),'فرم شماره 3- اطلاعات جامع نیروهای شرکتی')]").click()



companies = ['Kol']

# companies = ['SazMotori','SazMotori']
forms = ['Frm3']
i = 0
attachmentError = ''
for company in companies:
    wb = load_workbook(os.path.join(dirname,'delete.xlsx'))

    if('Frm3' in forms):
        ws = wb['Frm3']
        added = 0
        withError = 0
        col = []
        values = []

        index = 0
        total = ws.max_row
        for rownum in ws.iter_rows():
            index = index + 1
            try:
                if index == 1:
                    col = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]
                else:
                    values = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]

                    try:
                        webdriver.find_element_by_xpath("//label[contains(text(),'کد ملی')]/../input").clear()
                        webdriver.find_element_by_xpath("//label[contains(text(),'کد ملی')]/../input").send_keys(values[0])
                        webdriver.find_element_by_xpath("//label[contains(text(),'کد ملی')]/../input").send_keys(u'\ue007')

                        sleep(delaytime)

                        webdriver.switch_to_default_content()
                        webdriver.switch_to.frame(mainFrame)
                        msg = "پیدا نشد"
                        el = webdriver.find_element_by_xpath("//a/img[contains(@title,'حذف')]/..")
                        msg = "حذف نشد"
                        webdriver.execute_script("arguments[0].click();", el)
                        sleep(delaytime)
                        webdriver.switch_to.alert.accept(); 
                        msg = "موفق"
                        # webdriver.find_element_by_xpath().click()                                                             
                        print('{0} - {1} {2:3.0f}% '.format(index,total,index/total*100))
                    except:              
                        print('{0} - {1} {2:3.0f}% ERROR'.format(index,total,index/total*100)) 
                        # داشبورد مدیریت
                    ws['F'+str(index)] = msg
    
            except:
                print('{0} - {1} {2:3.0f}% ERROR'.format(index,total,index/total*100))                
                withError += 1
                continue
        try:
            wb.save(os.path.join(dirname, 'delete.xlsx'))
        except:
            print('file open')
        print('Frm3 {} Added {} Person.'.format(company,added))
        print('Frm3 {} With {} Error.'.format(company,withError))