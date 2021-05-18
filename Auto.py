from selenium import webdriver
from selenium.webdriver.common import keys
from selenium.webdriver.common.keys import Keys
from time import sleep, strftime
from random import randint
import pandas as pd
from openpyxl import load_workbook
import os
import sys
import jdatetime

dirname = 'D:\\Report\\1400\\TabdilVaz'
progDirname = os.path.dirname(__file__)
currentDirname = os.getcwd()
delaytime = 3

chromedriver_path = r'D://chromedriver/chromedriver.exe' # Change this to your own chromedriver path!
webdriver = webdriver.Chrome(executable_path=chromedriver_path)
# webdriver.set_window_size(1920, 1080)
webdriver.maximize_window()

def sendInput(id, txt):
    el = webdriver.find_element_by_xpath("/html/body/div/div/div/div/div/div/div/section[2]/form/fieldset/div/div/fieldset/div/div[1]/div["+str(id)+"]/input")
    el.clear()
    el.send_keys(txt)

def selectOption(id, val):
    el = webdriver.find_element_by_xpath("/html/body/div/div/div/div/div/div/div/section[2]/form/fieldset/div/div/fieldset/div/div[1]/div["+str(id)+"]/div/select/option[@value='"+str(val)+"']")
    el.click()

def sendDate(id, txt):
    if(len(txt) > 0):
        splitDay = txt.split("/")
        gregorian_date = jdatetime.date(int(splitDay[0]),int(splitDay[1]),int(splitDay[2])).togregorian()
        day = webdriver.find_element_by_xpath("/html/body/div/div/div/div/div/div/div/section[2]/form/fieldset/div/div/fieldset/div/div[1]/div["+str(id)+"]/span[1]/input[1]")
        month = webdriver.find_element_by_xpath("/html/body/div/div/div/div/div/div/div/section[2]/form/fieldset/div/div/fieldset/div/div[1]/div["+str(id)+"]/span[1]/input[2]")
        year = webdriver.find_element_by_xpath("/html/body/div/div/div/div/div/div/div/section[2]/form/fieldset/div/div/fieldset/div/div[1]/div["+str(id)+"]/span[1]/input[3]")
        webdriver.execute_script("arguments[0].setAttribute('value', '" + str(gregorian_date.day) +"')", day)
        webdriver.execute_script("arguments[0].setAttribute('value', '" + str(gregorian_date.month - 1) +"')", month)
        webdriver.execute_script("arguments[0].setAttribute('value', '" + str(gregorian_date.year) +"')", year)
def sendFile(id, txt):
    el = webdriver.find_element_by_xpath("/html/body/div/div/div/div/div/div/div/section[2]/form/fieldset/div/div/fieldset/div/div[1]/div["+str(id)+"]/input[1]")
    el.send_keys(txt)

def getFirstPage():
    try:
        # داشبورد مدیریت
        webdriver.get('https://amar.imo.org.ir/member/-_-irn-l-14000277232')
        webdriver.find_element_by_xpath("//*[contains(text(),'نمایش داده های آماری')]/..").click()
        sleep(delaytime)
        mainFrame = webdriver.find_element_by_xpath("//iframe")
        # webdriver.switch_to_frame('_cartable_WAR_Cartableportlet_dashboard-application_iframe_')
        webdriver.switch_to.frame(mainFrame)
        webdriver.find_element_by_xpath("//*[contains(text(),'مجموعه اطلاعات جامع اداری و استخدامی')]").click()
        webdriver.find_element_by_xpath("//*[contains(text(),'فرم شماره 3- اطلاعات جامع نیروهای شرکتی')]").click()
    except:
        getFirstPage()


def writeExcelCell(excelWB, index, cell1, cell2, cell3, cell4):
    ws['AM'+str(index)] = cell1
    ws['AO'+str(index)] = cell2
    ws['AN'+str(index)] = cell3
    ws['AP'+str(index)] = cell4

webdriver.get('https://amar.imo.org.ir/')
html = webdriver.page_source
userData = open("userData.udata", "r")

username = webdriver.find_element_by_name('_58_INSTANCE_dehyari_login')
username.send_keys(userData.readline().rstrip())
password = webdriver.find_element_by_name('_58_INSTANCE_dehyari_password')
password.send_keys(userData.readline().rstrip())

webdriver.find_element_by_xpath('//button[text()=" ورود "]').click()
companies = userData.readline().rstrip().split(',')
print(companies)
# companies = ['SazMotori','SazMotori']
forms =  userData.readline().rstrip().split(',')
excelfile =  userData.readline().rstrip()
if len(sys.argv) >= 2:
    excelfile = sys.argv[1]
if len(sys.argv) >= 3:
    forms = sys.argv[2].split(',')
print(forms)

companies = ['SH']
forms = ['Frm3']
i = 0

attachmentError = ''

getFirstPage()

for company in companies:
    wb = load_workbook(os.path.join(dirname, company+excelfile))

    if('Frm2' in forms):
        ws = wb['Frm2']

    if('Frm3' in forms):
        ws = wb['Frm3']
        added = 0
        withError = 0
        col = []
        values = []

        writeExcelCell(ws, 1, 'state', 'error', 'place', 'serial')
        # http://amarnameh.imo.org.ir/Input/Update.aspx?Id=8023&cid=281
        index = 0
        AddOrUpdate = ""
        total = ws.max_row
        for rownum in ws.iter_rows():
            index = index + 1
            try:
                if index == 1:
                    col = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]
                else:
                    values = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]

                    if(values[38] == 'added'):
                        continue
                    elif(values[41] == ''):
                        webdriver.find_element_by_xpath("//label[contains(text(),'کد ملی')]/../input").clear()
                        webdriver.find_element_by_xpath("//label[contains(text(),'کد ملی')]/../input").send_keys(values[4])
                        webdriver.find_element_by_xpath("//label[contains(text(),'کد ملی')]/../input").send_keys(u'\ue007')

                        sleep(delaytime)

                        AddOrUpdate = "Error finding"
                        try:
                            msg = "پیدا نشد"
                            el = webdriver.find_element_by_xpath("//a/i[contains(@title,'ویرایش')]/..")
                            msg = "ویرایش نشد"
                            webdriver.execute_script("arguments[0].click();", el)
                            AddOrUpdate = "update"
                        except:
                            webdriver.find_element_by_id("_srp_amar_amarmanagement_dynamic_dataviewer_WAR_AmarManagementportlet_add").click()
                            AddOrUpdate = "add"

                    sendInput(2,values[1])
                    sendInput(3,values[2])
                    sendInput(4,values[3])
                    sendInput(5,values[4])
                    selectOption(6,values[5])
                    selectOption(7,values[6])
                    if(values[6] == "237"):
                        sendInput(8,values[7])
                    selectOption(9,values[8])
                    sendDate(10,values[9])
                    selectOption(11,values[10])
                    if(values[10] == 1):
                        sendInput(12,values[11])
                    selectOption(13,values[12])                    
                    if(values[12] == 1):
                        selectOption(14,values[13])                    
                        sendDate(15,values[14])
                        sendDate(16,values[15])
                        sendDate(17,values[16])
                        sendInput(18,values[17])
                        sendDate(19,values[18])
                        sendDate(20,values[19])
                    sendDate(21,values[20])
                    sendDate(22,values[21])
                    sendInput(23,values[25])
                    sendInput(24,values[26])
                    sendInput(25,values[27])
                    sendInput(26,values[28])


                    sendInput(30,values[32])
                    sendInput(31,values[35])

                    msg = "موفق"                                                     

                    temp = os.path.join(progDirname,'1234567890.pdf')
                    attachmentError = ''
                    i=-1
                    attachmentFile = os.path.join(dirname,company+'\\'+values[4]+'t.pdf')
                    if(not(os.path.exists(attachmentFile))):
                        attachmentFile = temp
                        attachmentError = attachmentError + ' t'
                    else:
                        sendFile(27,attachmentFile)# مدرک تحصیلی

                    attachmentFile = os.path.join(dirname,company+'\\'+values[4]+'.pdf')
                    if(not(os.path.exists(attachmentFile))):
                        attachmentFile = temp
                        attachmentError = attachmentError + ' s'
                    else:
                        sendFile(28,attachmentFile)# سابقه بیمه       
                        
                    attachmentFile = os.path.join(dirname,company+'\\'+values[4]+'g.pdf')
                    if(not(os.path.exists(attachmentFile))):
                        attachmentFile = temp
                        attachmentError = attachmentError + ' g'
                    else:
                        sendFile(29,attachmentFile)# قرارداد        

                    if(AddOrUpdate == "add"):
                        webdriver.find_element_by_xpath('/html/body/div/div/div/div/div/div/div/section[2]/form/fieldset/div/div/fieldset/div/div[2]/button[2]').click()# finish     
                    else:
                        webdriver.find_element_by_xpath('/html/body/div/div/div/div/div/div/div/section[2]/form/fieldset/div/div/fieldset/div/div[3]/button[2]').click()# finish     
                    sleep(delaytime)
                    # 
                    # webdriver.find_element_by_xpath('/html/body/div/div/div/div/div/div/div/section[2]/form/div/div[2]/button').click()# finish     
                                                                    
                    # if(attachmentError == ''):
                    added += 1
                    writeExcelCell(ws, index, 'added', attachmentError, '', '')
                    print('{0} Frm3 : {1} - {2} {3:3.0f}% Adeed {4} '.format(company,index,total,index/total*100,values[4]))

                    # else:
                    #     withError += 1
                    #     writeExcelCell(ws, index, "error", 'attachment', attachmentError, '')
                    #     print('{0} Frm3 : {1} - {2} {3:3.0f}% Adeed AttachmentError {4} - {5}'.format(company,index,total,index/total*100,values[4], attachmentError))
                   
                    getFirstPage()
                    try:
                        wb.save(os.path.join(dirname, company+excelfile))
                    except:
                        print('file open') 
                   
            except:
                writeExcelCell(ws, index, "error", col[i], values[i], '')
                withError += 1
                print('{0} Frm3 : {1} - {2} {3:3.0f}%  Error {4} - {5} {6} {7}'.format(company,index,total,index/total*100,values[9], i,col[i], values[i]))                                 
                
                getFirstPage()

                continue

        try:
            wb.save(os.path.join(dirname, company+excelfile))
        except:
            print('file open') 

        print('Frm3 {} Added {} Person.'.format(company,added))
        print('Frm3 {} With {} Error.'.format(company,withError))
webdriver.stop_client()
webdriver.close()