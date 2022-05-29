from ast import While
from fileinput import filename
from selenium import webdriver
from selenium.webdriver.common import keys
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

from time import sleep, strftime
from random import randint
import pandas as pd
from openpyxl import load_workbook
import os
import sys
import jdatetime

dirname = 'D:\\Report\\1401\\Taavoni'
file_name = 'All.xlsx'
progDirname = os.path.dirname(__file__)
currentDirname = os.getcwd()
delaytime = 1

chromedriver_path = r'D://chromedriver/chromedriver.exe' # Change this to your own chromedriver path!
webdriver = webdriver.Chrome(executable_path=chromedriver_path)
# webdriver.set_window_size(1920, 1080)
webdriver.maximize_window()

def getFirstPage():
    try:
        # داشبورد مدیریت
        webdriver.get('https://taavoni.mcls.gov.ir/addUserToCompany.php')

        webdriver.find_element(By.ID,"nationalID").send_keys("14008213497")
        webdriver.find_element(By.XPATH,"//*[@id=\"companyFormElm\"]/div/div/a").click()

        webdriver.find_element(By.ID,"postalCode").send_keys("8136613348")
        webdriver.execute_script('document.getElementById("address").removeAttribute("readonly")')
        webdriver.execute_script('document.getElementById("address2").removeAttribute("readonly")')

        webdriver.find_element(By.ID,"address").send_keys("ستان اصفهان ، شهرستان اصفهان ، بخش مرکزی ، شهر اصفهان، محله درب کوشک ، خیابان باب الرحمه ، میدان امام حسین ، پلاک 2 ، ساختمان ارگ جهان نما ، فاز3 ، طبقه دوم ، واحد 24")
        webdriver.find_element(By.ID,"address2").send_keys("ستان اصفهان ، شهرستان اصفهان ، بخش مرکزی ، شهر اصفهان، محله درب کوشک ، خیابان باب الرحمه ، میدان امام حسین ، پلاک 2 ، ساختمان ارگ جهان نما ، فاز3 ، طبقه دوم ، واحد 24")

        webdriver.find_element(By.XPATH,"//*[@id=\"requestType\"]/option[3]").click()
        webdriver.find_element(By.XPATH,"//*[@id=\"memberType\"]/option[2]").click()        
    except:
        getFirstPage()


def writeExcelCell(excelWB, index, cell1, cell2, cell3, cell4):
    ws['X'+str(index)] = cell1
    ws['Y'+str(index)] = cell2
    ws['Z'+str(index)] = cell3
    ws['AA'+str(index)] = cell4

i = 0

wb = load_workbook(os.path.join(dirname, file_name))
ws = wb['All']

added = 0
withError = 0
col = []
values = []

writeExcelCell(ws, 1, 'state', 'error', 'place', 'msg')
index = 0
AddOrUpdate = ""
total = ws.max_row
for rownum in ws.iter_rows():
    index = index + 1
    try:
        if index == 1:
            col = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]
        else:
            values = [(u"" if cell.value is None else str(cell.value).strip()) for cell in rownum]

            if(values[23] == 'added'):
                continue
            else:

                webdriver.find_element(By.ID,"nationalCode").send_keys(values[1])

                if values[10] == "مرد":
                    webdriver.find_element(By.XPATH,"//*[@id=\"gender\"]/option[2]").click()
                else:
                    webdriver.find_element(By.XPATH,"//*[@id=\"gender\"]/option[3]").click()

                webdriver.find_element(By.ID,"name").send_keys(values[3])

                webdriver.find_element(By.ID,"family").send_keys(values[4])

                webdriver.find_element(By.ID,"fatherName").send_keys(values[14])

                webdriver.find_element(By.ID,"birthDay").send_keys(values[12])

                webdriver.find_element(By.ID,"mobile").send_keys(values[9])

                webdriver.find_element(By.XPATH,"//*[@id=\"position\"]/option[2]").click()

                webdriver.find_element(By.ID,"sendForm").click()
                                                                

                msg = webdriver.find_element(By.ID,"message").get_attribute('innerHTML')
                while msg == "":
                    sleep(delaytime)
                    msg = webdriver.find_element(By.ID,"message").get_attribute('innerHTML')
 
                writeExcelCell(ws, index, "added", "", "",msg)
                
                try:
                    print('{0} - {1} {2:3.0f}%  Ok {3} '.format(index,total,index/total*100,values[1]))                                                 
                    wb.save(os.path.join(dirname, file_name))
                except:
                    print('file open') 
        getFirstPage()
    except:
        writeExcelCell(ws, index, "error", col[i], values[i], '')
        withError += 1
        print('{0} - {1} {2:3.0f}%  Error {3} - {4} {5} {6}'.format(index,total,index/total*100,values[1], i,col[i], values[i]))                                                 
        getFirstPage()

    continue

try:
    wb.save(os.path.join(dirname, file_name))
except:
    print('file open')
    input("Press Enter to continue...")
    wb.save(os.path.join(dirname, file_name))

webdriver.stop_client()
webdriver.close()